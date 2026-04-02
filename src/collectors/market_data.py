"""Import bank financial data from Excel (bb_values.xlsx format)."""

import logging
from datetime import datetime
from pathlib import Path

import openpyxl

from src.collectors.base import BaseCollector
from src.models.database import Database

logger = logging.getLogger(__name__)

# Map Excel bank abbreviations to competitor IDs
BANK_ID_MAP = {
    "RB": "raiffeisenbank",
    "RB ": "raiffeisenbank",
    "Moneta": "moneta",
    "ČSOB": "csob",
    "ČS": "ceska_sporitelna",
    "ČS ": "ceska_sporitelna",
    "KB": "komercni_banka",
    "FIO": "fio_banka",
    "UNI": "unicredit",
    "Air Bank": "air_bank",
    "AIR": "air_bank",
    "Partners": "partners_bank",
    "PB": "partners_bank",
    "Revolut": "revolut_cz",
    "REV": "revolut_cz",
}

# ── Sheet 1: values ─────────────────────────────────────────────

METRIC_MAP = {
    "Total Assets": ("total_assets", "balance_sheet", "mio CZK"),
    "Total Liabilities": ("total_liabilities", "balance_sheet", "mio CZK"),
    "Total Equity": ("total_equity", "balance_sheet", "mio CZK"),
    "Loans and receivables to Banks": ("loans_to_banks", "balance_sheet", "mio CZK"),
    "Deposits received from Banks": ("deposits_from_banks", "balance_sheet", "mio CZK"),
    "Loans and receivables to Customers": ("loans_to_customers", "loans", "mio CZK"),
    "Deposits received from Customers": ("customer_deposits", "balance_sheet", "mio CZK"),
    "Intangible Assets": ("intangible_assets", "balance_sheet", "mio CZK"),
    "Debt Securties": ("debt_securities", "balance_sheet", "mio CZK"),
    "ROE - QTD": ("roe_qtd", "profitability", "%"),
    "ROE - YTD": ("roe_ytd", "profitability", "%"),
    "Operating Income (mio CZK) - QTD": ("op_income_qtd", "income", "mio CZK"),
    "Operating Income (mio CZK) - YTD": ("op_income_ytd", "income", "mio CZK"),
    "Operating Expense (mio CZK) - QTD": ("op_expense_qtd", "expenses", "mio CZK"),
    "Operating Expense (mio CZK) - YTD": ("op_expense_ytd", "expenses", "mio CZK"),
    "Net Operating Income (mio CZK) - QTD": ("net_op_income_qtd", "income", "mio CZK"),
    "Net Operating Income (mio CZK) - YTD": ("net_op_income_ytd", "income", "mio CZK"),
    "Cost to Income Ratio (%) - QTD": ("cir_qtd", "profitability", "%"),
    "Cost to Income Ratio (%) - YTD": ("cir_ytd", "profitability", "%"),
    "Net Profit after tax (mio CZK) - QTD": ("npat_qtd", "profitability", "mio CZK"),
    "Net Profit after tax (mio CZK) - YTD": ("npat_ytd", "profitability", "mio CZK"),
    "Profit after tax  per Employee (CZK) - QTD": ("npat_per_employee_qtd", "efficiency", "CZK"),
    "Profit after tax per Client (CZK) - QTD": ("npat_per_client_qtd", "efficiency", "CZK"),
    "Operating Income per Employee (CZK) - QTD": ("op_income_per_employee_qtd", "efficiency", "CZK"),
    "Operating Income per Client (CZK) - QTD": ("op_income_per_client_qtd", "efficiency", "CZK"),
    "Operating Expense per Employee (CZK) - QTD": ("op_expense_per_employee_qtd", "efficiency", "CZK"),
    "Operating Expense per Client (CZK) - QTD": ("op_expense_per_client_qtd", "efficiency", "CZK"),
    "Net Operating Income per Employee (CZK) - QTD": ("net_op_income_per_employee_qtd", "efficiency", "CZK"),
    "Net Operating Income per Client (CZK) - QTD": ("net_op_income_per_client_qtd", "efficiency", "CZK"),
    "Number of FTE - EoP": ("fte", "operations", "FTE"),
    "Number of clients (mio) - EoP": ("clients", "operations", "mil"),
    "Loans Total (mio CZK)": ("loans_total", "loans", "mio CZK"),
    "Retail Loans Total (mio CZK)": ("loans_retail", "loans", "mio CZK"),
    "Commercial Loans Total (mio CZK)": ("loans_commercial", "loans", "mio CZK"),
    "Mortgages Total (mio CZK)": ("mortgages", "loans", "mio CZK"),
    "Interest Income (mio CZK) - QTD": ("interest_income_qtd", "nii", "mio CZK"),
    "Interest Income (mio CZK) - YTD": ("interest_income_ytd", "nii", "mio CZK"),
    "Interest Expense (mio CZK) - QTD": ("interest_expense_qtd", "nii", "mio CZK"),
    "Interest Expense (mio CZK) - YTD": ("interest_expense_ytd", "nii", "mio CZK"),
    "Net Interest Income (mio CZK) - QTD": ("nii_qtd", "nii", "mio CZK"),
    "Net Interest Income (mio CZK) - YTD": ("nii_ytd", "nii", "mio CZK"),
    "Net Interest Margin - QTD": ("nim_qtd", "nii", "%"),
    "Net Interest Margin - YTD": ("nim_ytd", "nii", "%"),
    "Net Fees&Commissions  (mio CZK) - QTD": ("net_fees_qtd", "income", "mio CZK"),
    "Net Fees&Commissions  (mio CZK) - YTD": ("net_fees_ytd", "income", "mio CZK"),
    "Personnel Expense (mio CZK) - QTD": ("perex_qtd", "expenses", "mio CZK"),
    "Personnel Expense (mio CZK) - YTD": ("perex_ytd", "expenses", "mio CZK"),
    "General Administrative Expenses (mio CZK) - QTD": ("gae_qtd", "expenses", "mio CZK"),
    "General Administrative Expenses (mio CZK) - YTD": ("gae_ytd", "expenses", "mio CZK"),
    "Regulatory charges (mio CZK) - QTD": ("reg_charges_qtd", "expenses", "mio CZK"),
    "Regulatory charges (mio CZK) - YTD": ("reg_charges_ytd", "expenses", "mio CZK"),
    "Depreciation and Amortisation (mio CZK) - QTD": ("depreciation_qtd", "expenses", "mio CZK"),
    "Depreciation and Amortisation (mio CZK) - YTD": ("depreciation_ytd", "expenses", "mio CZK"),
    "Other Operating Result (mio CZK) - QTD": ("other_op_result_qtd", "income", "mio CZK"),
    "Other Operating Result (mio CZK) - YTD": ("other_op_result_ytd", "income", "mio CZK"),
    "Personnel Expense per Employee (CZK) - QTD": ("perex_per_employee_qtd", "efficiency", "CZK"),
    "GAE per Employee (CZK) - QTD": ("gae_per_employee_qtd", "efficiency", "CZK"),
    "Personnel Expense per Client (CZK) - QTD": ("perex_per_client_qtd", "efficiency", "CZK"),
    "GAE per Client(CZK) - QTD": ("gae_per_client_qtd", "efficiency", "CZK"),
    "Liquidity Coverage Ratio": ("lcr", "regulatory", "%"),
    "Net Stable Funding Ratio": ("nsfr", "regulatory", "%"),
    "Effective Tax Rate - YTD": ("effective_tax_rate_ytd", "profitability", "%"),
    "Capital Adequacy (% of RWA)": ("capital_adequacy", "regulatory", "%"),
    "Non-Performing Loans (%)": ("npl_ratio", "risk", "%"),
    "MREL": ("mrel", "regulatory", "%"),
    "Ratings (long-term)": ("rating", "regulatory", "rating"),
    "Risk Costs - QTD (mio CZK)": ("risk_costs_qtd", "risk", "mio CZK"),
    "Risk Weighted Assets (mio CZK)": ("rwa", "risk", "mio CZK"),
    "Risk Charge (%)": ("risk_charge", "risk", "%"),
}


def import_financials(db: Database, excel_path: str) -> int:
    """Parse bb_values.xlsx and load all sheets into the DB."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    total = 0

    # Sheet 1: values (main comparison data)
    if "values" in wb.sheetnames:
        total += _import_values_sheet(db, wb["values"])
    elif "Sheet1" in wb.sheetnames:
        total += _import_values_sheet(db, wb["Sheet1"])

    # Sheet 2: loan_drill_down
    if "loan_drill_down" in wb.sheetnames:
        total += _import_loan_drilldown(db, wb["loan_drill_down"])

    # Sheet 3: detailed_values — skipped (different reporting entities than values sheet)

    wb.close()
    logger.info("Total imported: %d data points", total)
    return total


# ── Sheet 1: values ─────────────────────────────────────────────

def _import_values_sheet(db: Database, ws) -> int:
    total = 0
    row = 1
    max_row = ws.max_row

    while row <= max_row:
        metric_row = None
        for r in range(row, min(row + 5, max_row + 1)):
            cell_b = ws.cell(row=r, column=2).value
            if cell_b and _is_date(cell_b):
                metric_row = r
                break

        if metric_row is None:
            row += 1
            continue

        series_name = str(ws.cell(row=metric_row, column=1).value or "").strip()
        if not series_name:
            row = metric_row + 10
            continue

        series_id, category, unit = _classify_metric(series_name)

        dates = []
        for col in range(2, ws.max_column + 1):
            val = ws.cell(row=metric_row, column=col).value
            if val and _is_date(val):
                dates.append((col, _parse_date(val)))
            elif val is None:
                break

        if not dates:
            row = metric_row + 10
            continue

        data_start = metric_row + 2
        for r in range(data_start, min(data_start + 8, max_row + 1)):
            bank_label = ws.cell(row=r, column=1).value
            if not bank_label:
                continue
            bank_label = str(bank_label).strip()
            comp_id = BANK_ID_MAP.get(bank_label)
            if not comp_id:
                continue

            for col, date_str in dates:
                val = ws.cell(row=r, column=col).value
                if val is None:
                    continue
                try:
                    value = float(val)
                except (ValueError, TypeError):
                    continue
                db.upsert_metric("excel", series_id, series_name, category, date_str, value, unit, comp_id)
                total += 1

        row = data_start + 8

    logger.info("values sheet: %d data points", total)
    return total


# ── Sheet 2: loan_drill_down ────────────────────────────────────

def _import_loan_drilldown(db: Database, ws) -> int:
    total = 0

    # Part A: horizontal blocks (cols 1-63) — Loans Total, Retail, Commercial, Mortgages by bank
    # Already captured in values sheet, but also has Q/Q and Y/Y changes
    # We'll import the Q/Q % changes as separate series
    blocks = [
        (1, "loans_total", "Loans Total"),
        (17, "loans_retail", "Retail Loans Total"),
        (33, "loans_commercial", "Commercial Loans Total"),
        (49, "mortgages", "Mortgages Total"),
    ]

    for start_col, base_id, base_name in blocks:
        # Row 17 onwards has Q/Q Change (%) - import those
        dates = []
        for col in range(start_col + 1, start_col + 15):
            val = ws.cell(row=1, column=col).value
            if val and _is_date(val):
                dates.append((col, _parse_date(val)))

        # Q/Q Change % (row 17-23 area)
        for r in range(17, 24):
            bank_label = ws.cell(row=r, column=start_col).value
            if not bank_label:
                continue
            comp_id = BANK_ID_MAP.get(str(bank_label).strip())
            if not comp_id:
                continue
            for col, date_str in dates:
                val = ws.cell(row=r, column=col).value
                if val is None:
                    continue
                try:
                    value = float(val)
                except (ValueError, TypeError):
                    continue
                sid = f"{base_id}_qoq_pct"
                db.upsert_metric("excel", sid, f"{base_name} Q/Q Change %", "loans_growth", date_str, value, "%", comp_id)
                total += 1

        # Y/Y Change % (row 25-31 area)
        for r in range(25, 32):
            bank_label = ws.cell(row=r, column=start_col).value
            if not bank_label:
                continue
            comp_id = BANK_ID_MAP.get(str(bank_label).strip())
            if not comp_id:
                continue
            for col, date_str in dates:
                val = ws.cell(row=r, column=col).value
                if val is None:
                    continue
                try:
                    value = float(val)
                except (ValueError, TypeError):
                    continue
                sid = f"{base_id}_yoy_pct"
                db.upsert_metric("excel", sid, f"{base_name} Y/Y Change %", "loans_growth", date_str, value, "%", comp_id)
                total += 1

    # Part B skipped — drill-down computed server-side from Retail/Commercial/Mortgages/Other

    logger.info("loan_drill_down sheet: %d data points", total)
    return total


def _import_bank_loan_breakdown(db: Database, ws, start_col: int, unit: str) -> int:
    """Parse the per-bank loan breakdown sections starting at start_col."""
    total = 0
    row = 1
    max_row = ws.max_row

    while row <= max_row:
        # Look for bank header (e.g. "Moneta (mio CZK)" or "Loan Structure - Moneta")
        header = ws.cell(row=row, column=start_col).value
        if not header:
            row += 1
            continue

        header = str(header).strip()

        # Identify which bank
        comp_id = None
        for bank_key, cid in BANK_ID_MAP.items():
            if bank_key.strip() in header:
                comp_id = cid
                break

        if not comp_id:
            row += 1
            continue

        # Parse dates from this header row
        dates = []
        for col in range(start_col + 1, start_col + 15):
            val = ws.cell(row=row, column=col).value
            if val and (_is_date(val) or _is_quarter_label(val)):
                dates.append((col, _parse_quarter_or_date(val)))

        if not dates:
            row += 1
            continue

        is_pct = "Structure" in header or unit == "%"
        suffix = "_pct" if is_pct else ""

        # Read data rows until empty row or next bank header
        r = row + 1
        while r <= max_row:
            label = ws.cell(row=r, column=start_col).value
            if not label:
                break
            label = str(label).strip()
            if any(bk in label for bk in ["(mio CZK)", "Structure -"]):
                break  # next bank section

            # Create series_id from the loan category
            slug = label.lower().replace(" ", "_").replace("&", "and").replace("+", "_")
            slug = slug.replace("(", "").replace(")", "").replace(",", "").replace("/", "_")[:40]
            series_id = f"loan_{comp_id}_{slug}{suffix}"
            series_name = f"{label}{' (%)' if is_pct else ''}"

            for col, date_str in dates:
                val = ws.cell(row=r, column=col).value
                if val is None:
                    continue
                try:
                    value = float(val)
                except (ValueError, TypeError):
                    continue
                db.upsert_metric("excel", series_id, series_name, "loan_drilldown", date_str, value, unit, comp_id)
                total += 1
            r += 1

        row = r

    return total


# ── Sheet 3: detailed_values ────────────────────────────────────

DETAILED_ROW_MAP = {
    "Assets": ("detailed_assets", "balance_sheet", "mio CZK"),
    "Customer Loans": ("detailed_customer_loans", "loans", "mio CZK"),
    "Liabilities": ("detailed_liabilities", "balance_sheet", "mio CZK"),
    "Customer Deposits": ("detailed_customer_deposits", "balance_sheet", "mio CZK"),
    "Equity": ("detailed_equity", "balance_sheet", "mio CZK"),
    "RWA": ("detailed_rwa", "risk", "mio CZK"),
    "Gross Income": ("detailed_gross_income", "income", "mio CZK"),
    "NII": ("detailed_nii", "nii", "mio CZK"),
    "NCFI": ("detailed_ncfi", "income", "mio CZK"),
    "Other": ("detailed_other_income", "income", "mio CZK"),
    "OPEX": ("detailed_opex", "expenses", "mio CZK"),
    "PEREX": ("detailed_perex", "expenses", "mio CZK"),
    "GAE": ("detailed_gae", "expenses", "mio CZK"),
    "D&A": ("detailed_da", "expenses", "mio CZK"),
    "DIC + RF": ("detailed_dic_rf", "expenses", "mio CZK"),
    "Risk Costs": ("detailed_risk_costs", "risk", "mio CZK"),
    "Profit Before Tax": ("detailed_pbt", "profitability", "mio CZK"),
    "Tax": ("detailed_tax", "profitability", "mio CZK"),
    "NPAT": ("detailed_npat", "profitability", "mio CZK"),
    "ROE": ("detailed_roe", "profitability", "%"),
    "CIR": ("detailed_cir", "profitability", "%"),
    "CIR (OPEX/Oper Income)": ("detailed_cir_opex_oi", "profitability", "%"),
}


def _import_detailed_values(db: Database, ws) -> int:
    total = 0
    row = 1
    max_row = ws.max_row

    while row <= max_row:
        # Look for bank header row (col B contains "Bank (mio CZK)" and col C has a date)
        header = ws.cell(row=row, column=2).value
        if not header or not isinstance(header, str):
            row += 1
            continue

        header = header.strip()
        comp_id = None
        for bank_key, cid in BANK_ID_MAP.items():
            if bank_key.strip() in header:
                comp_id = cid
                break

        if not comp_id:
            row += 1
            continue

        # Parse dates from this header row (col C onwards)
        dates = []
        for col in range(3, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val and _is_date(val):
                dates.append((col, _parse_date(val)))

        if not dates:
            row += 1
            continue

        # Read data rows
        r = row + 1
        while r <= max_row:
            label = ws.cell(row=r, column=2).value
            if not label:
                r += 1
                # Check if we hit the next bank section
                next_header = ws.cell(row=r, column=2).value
                if next_header and isinstance(next_header, str) and "(mio CZK)" in next_header:
                    break
                continue

            label = str(label).strip()

            # Skip sub-labels like "(OPEX/GI)"
            if label.startswith("("):
                r += 1
                continue

            mapping = DETAILED_ROW_MAP.get(label)
            if not mapping:
                r += 1
                continue

            series_id, category, unit = mapping

            for col, date_str in dates:
                val = ws.cell(row=r, column=col).value
                if val is None:
                    continue
                try:
                    value = float(val)
                except (ValueError, TypeError):
                    continue
                db.upsert_metric("excel", series_id, f"{label} (QTD)", f"detailed_{category}", date_str, value, unit, comp_id)
                total += 1

            r += 1

        row = r

    logger.info("detailed_values sheet: %d data points", total)
    return total


# ── Helpers ─────────────────────────────────────────────────────

def _is_date(val) -> bool:
    if isinstance(val, datetime):
        return True
    s = str(val).strip()
    return "/" in s and len(s) <= 12


def _is_quarter_label(val) -> bool:
    s = str(val).strip()
    return s[:2] in ("1Q", "2Q", "3Q", "4Q") and len(s) >= 7


def _parse_date(val) -> str:
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    parts = s.split("/")
    if len(parts) == 3:
        month, day, year = parts
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    return s


def _parse_quarter_or_date(val) -> str:
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # "1Q 2022" -> "2022-03-31"
    quarter_end = {"1Q": "03-31", "2Q": "06-30", "3Q": "09-30", "4Q": "12-31"}
    for q, end in quarter_end.items():
        if s.startswith(q):
            year = s.split()[-1] if " " in s else s[2:]
            return f"{year}-{end}"
    return _parse_date(val)


def _classify_metric(name: str) -> tuple[str, str, str]:
    if name in METRIC_MAP:
        return METRIC_MAP[name]
    for key, (sid, cat, unit) in METRIC_MAP.items():
        if key in name or name in key:
            return sid, cat, unit
    slug = name.lower().replace(" ", "_").replace("(", "").replace(")", "")[:40]
    return slug, "other", ""


class MarketDataCollector(BaseCollector):
    """Collector that imports bank financials from Excel."""
    name = "market_data"
    rate_limit_delay = 0
    required_source_key = None

    def collect(self, competitor_id: str) -> list:
        return []

    def run(self, competitor_ids: list[str] | None = None) -> dict:
        run_id = self.db.start_collector_run(self.name, None)

        excel_path = self.config_dir / "bb_values.xlsx"
        if not excel_path.exists():
            excel_path = Path("data") / "bb_values.xlsx"

        if not excel_path.exists():
            self.db.finish_collector_run(run_id, "failed", error_message="bb_values.xlsx not found")
            return {"total": 0, "new_signals": 0, "errors": 1,
                    "competitors": {"_market_data": {"status": "failed", "error": "bb_values.xlsx not found"}}}

        try:
            total = import_financials(self.db, str(excel_path))
            self.db.finish_collector_run(run_id, "success", total)
            return {"total": total, "new_signals": total, "errors": 0,
                    "competitors": {"_market_data": {"status": "success", "signals": total}}}
        except Exception as e:
            self.db.finish_collector_run(run_id, "failed", error_message=str(e))
            return {"total": 0, "new_signals": 0, "errors": 1,
                    "competitors": {"_market_data": {"status": "failed", "error": str(e)}}}
